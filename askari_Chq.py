import PIL
from PIL import ImageDraw, Image, ImageOps
import textwrap
import openpyxl as xl
from datetime import datetime
from amount_to_million import amount_to_million
import win32print
import win32ui
from PIL import Image, ImageWin
import os
import re

config = os.path.exists("askariConfig.ini")

def listToStringWithoutBrackets(value):
    return str(value).replace('[', '').replace(']', '').replace('(', '').replace(')', '').replace(',', '').replace(
        "'", '').replace(' ', '')

def askari_Chq(value):
    now = datetime.now()
    dates = str(now.strftime("%d-%m-%Y"))
    datex = dates.replace("-", "")
    datez = datex[:-2]

    chequeWidth = 811
    chequeHeight = 278

    chequeFont = 'arial.ttf'
    chequeFontSize = 20

    chequeDatex = 650
    chequeDatey = 60.8

    chequeNamex = 223
    chequeNamey = 99.28

    chequeAmountEnglishx = 248
    chequeAmountEnglishy = 125.2

    chequeAmountNumberx = 680
    chequeAmountNumbery = 122.4

    chequeFont = 'arial.ttf'
    chequeFontSize = 12

    counterDatex = 80.4
    counterDatey = 86.2

    counterNamex = 76.0
    counterNamey = 99.8

    counterAmountx = 95
    counterAmounty = 191.6

    if config == True:
        with open('askariConfig.ini') as f:
            for line in f:
                # For Python3, use print(line)
                if 'counterAmountx' in line:
                    txt = str(line)
                    receipts = str(re.findall(r'\d+(?:\.\d+)?', txt))
                    game = listToStringWithoutBrackets(receipts)
                    counterAmountx = float(game)
                if 'counterAmounty' in line:
                    txt = str(line)
                    receipts = str(re.findall(r'\d+(?:\.\d+)?', txt))
                    game = listToStringWithoutBrackets(receipts)
                    counterAmounty = float(game)
                if 'counterNamex' in line:
                    txt = str(line)
                    receipts = str(re.findall(r'\d+(?:\.\d+)?', txt))
                    game = listToStringWithoutBrackets(receipts)
                    counterNamex = float(game)
                if 'counterNamey' in line:
                    txt = str(line)
                    receipts = str(re.findall(r'\d+(?:\.\d+)?', txt))
                    game = listToStringWithoutBrackets(receipts)
                    counterNamey = float(game)
                if 'counterDatex' in line:
                    txt = str(line)
                    receipts = str(re.findall(r'\d+(?:\.\d+)?', txt))
                    game = listToStringWithoutBrackets(receipts)
                    counterDatex = float(game)
                if 'counterDatey' in line:
                    txt = str(line)
                    receipts = str(re.findall(r'\d+(?:\.\d+)?', txt))
                    game = listToStringWithoutBrackets(receipts)
                    counterDatey = float(game)
                if 'chequeAmountNumberx' in line:
                    txt = str(line)
                    receipts = str(re.findall(r'\d+(?:\.\d+)?', txt))
                    game = listToStringWithoutBrackets(receipts)
                    chequeAmountNumberx = float(game)
                if 'chequeAmountNumbery' in line:
                    txt = str(line)
                    receipts = str(re.findall(r'\d+(?:\.\d+)?', txt))
                    game = listToStringWithoutBrackets(receipts)
                    chequeAmountNumbery = float(game)
                if 'chequeAmountEnglishx' in line:
                    txt = str(line)
                    receipts = str(re.findall(r'\d+(?:\.\d+)?', txt))
                    game = listToStringWithoutBrackets(receipts)
                    chequeAmountEnglishx = float(game)
                if 'chequeAmountEnglishy' in line:
                    txt = str(line)
                    receipts = str(re.findall(r'\d+(?:\.\d+)?', txt))
                    game = listToStringWithoutBrackets(receipts)
                    chequeAmountEnglishy = float(game)
                if 'chequeNamex' in line:
                    txt = str(line)
                    receipts = str(re.findall(r'\d+(?:\.\d+)?', txt))
                    game = listToStringWithoutBrackets(receipts)
                    chequeNamex = float(game)
                if 'chequeNamey' in line:
                    txt = str(line)
                    receipts = str(re.findall(r'\d+(?:\.\d+)?', txt))
                    game = listToStringWithoutBrackets(receipts)
                    chequeNamey = float(game)
                if 'chequeDatex' in line:
                    txt = str(line)
                    receipts = str(re.findall(r'\d+(?:\.\d+)?', txt))
                    game = listToStringWithoutBrackets(receipts)
                    chequeDatey = float(game)
                if 'chequeDatey' in line:
                    txt = str(line)
                    receipts = str(re.findall(r'\d+(?:\.\d+)?', txt))
                    game = listToStringWithoutBrackets(receipts)
                    chequeDatey = float(game)

    wb = xl.load_workbook(value, data_only=True)
    # wb = xl.load_workbook(r'C:\Users\HYSTOU\Desktop\sest.xlsx')
    sheet = wb['Sheet1']
    cell = sheet.cell(1, 1)

    # creating image object which is of specific color
    width = 811
    height = 278

    black = 'black'
    white = 'white'

    for row in range(1, sheet.max_row + 1):
        img = PIL.Image.new(mode="RGB", size=(width, height),
                            color=(255, 255, 255))
        # img = ImageOps.expand(img, border=1, fill=black)
        draw = ImageDraw.Draw(img)
        draw.fontmode = "1"

        name = sheet.cell(row, 1)
        amount = sheet.cell(row, 2)

        # draw.line((1, 278) + img.size, fill=(0,0,0))

        # Cheque Filling

        font = PIL.ImageFont.truetype('arial.ttf', 16)
        draw.fontmode = "1"

        date = "  ".join(datez)

        draw.text((chequeDatex, chequeDatey), date + ' ' + datex[-4] + ' ' + datex[-1:], font=font, fill=black)
        draw.text((chequeNamex, chequeNamey), name.value, font=font, fill=black)
        fontSize = 16
        storeFontSize = fontSize
        wrapperWidth = 50
        caption = amount_to_million(amount.value) + " Rupees Only"

        length = len(str(caption))
        limit = 110

        if length > limit:
            calculationFontSize = int((limit * fontSize) / length)
            fontSize = calculationFontSize
            calculationWrapperWidth = (wrapperWidth * storeFontSize) / fontSize
            wrapperWidth = int(calculationWrapperWidth)



        caption1 = str(caption[0:wrapperWidth])
        caption2 = str(caption[wrapperWidth:])

        # print(caption1 + "\n" + caption2)
        count = int(wrapperWidth)

        for wrap in reversed(caption1):
            count -= 1
            if wrap == ' ' or wrap == '-':
                manualTextWrap = count + 1
                caption1 = str(caption[0:manualTextWrap])
                caption2 = str(caption[manualTextWrap:])
                break



        # Ratio Calculation: Will trigger resizing if the size of the value translated into words cross the lenght equal to 90
        # will reset the values of (1- Width For Text Wrap, 2- Font Size)


        font = PIL.ImageFont.truetype("arial.ttf", fontSize)
        # caption = caption
        # wrapper = textwrap.TextWrapper(width=wrapperWidth)
        # word_list = wrapper.wrap(text=caption)
        # caption_new = ''
        # for ii in word_list[:-1]:
        #     caption_new = caption_new + ii + '\n'

        # caption_new += word_list[-1]
        if len(str(caption)) > wrapperWidth:
            draw.text((chequeAmountEnglishx, chequeAmountEnglishy), caption1, font=font, fill=black)
            draw.text((chequeAmountEnglishx - 55, chequeAmountEnglishy + 23), caption2, font=font, fill=black)
        else:
            draw.text((chequeAmountEnglishx, chequeAmountEnglishy), caption, font=font, fill=black)
        font = PIL.ImageFont.truetype('arial.ttf', 16)
        payment = "{:,}".format(amount.value)
        draw.text((chequeAmountNumberx, chequeAmountNumbery), str(payment) + '/-', font=font, fill=black)

        # Counter Filling

        font = PIL.ImageFont.truetype("arial.ttf", 12)
        draw.fontmode = "1"
        draw.text((counterDatex, counterDatey), str(dates), font=font, fill=black)

        caption = name.value

        wrapper = textwrap.TextWrapper(width=16)
        word_list = wrapper.wrap(text=caption)
        caption_new = ''
        for ii in word_list[:-1]:
            caption_new = caption_new + ii + '\n'
        caption_new += word_list[-1]
        draw.text((counterNamex, counterNamey), caption_new, font=font, fill=black)
        draw.text((counterAmountx, counterAmounty), str(payment) + "/-", font=font, fill=black)

        # this will show image in any image viewer
        # img.show()
        img.save(name.value + ".jpg")


        # Constants for GetDeviceCaps
        #
        #
        # HORZRES / VERTRES = printable area

        HORZRES = 8
        VERTRES = 10
        #
        # LOGPIXELS = dots per inch
        #
        LOGPIXELSX = 88
        LOGPIXELSY = 90
        #
        # PHYSICALWIDTH/HEIGHT = total area
        #
        PHYSICALWIDTH = 110
        PHYSICALHEIGHT = 111
        #
        # PHYSICALOFFSETX/Y = left / top margin
        #
        PHYSICALOFFSETX = 112
        PHYSICALOFFSETY = 113

        printer_name = win32print.GetDefaultPrinter()
        file_name = "test.jpg"

        #
        # You can only write a Device-independent bitmap
        #  directly to a Windows device context; therefore
        #  we need (for ease) to use the Python Imaging
        #  Library to manipulate the image.
        #
        # Create a device context from a named printer
        #  and assess the printable size of the paper.
        #
        hDC = win32ui.CreateDC()
        hDC.CreatePrinterDC(printer_name)
        printable_area = hDC.GetDeviceCaps(HORZRES), hDC.GetDeviceCaps(VERTRES)
        printer_size = hDC.GetDeviceCaps(PHYSICALWIDTH), hDC.GetDeviceCaps(PHYSICALHEIGHT)
        printer_margins = hDC.GetDeviceCaps(PHYSICALOFFSETX), hDC.GetDeviceCaps(PHYSICALOFFSETY)

        #
        # Open the image, rotate it if it's wider than
        #  it is high, and work out how much to multiply
        #  each pixel by to get it as big as possible on
        #  the page without distorting.
        #
        bmp = Image.open(name.value + ".jpg")
        if bmp.size[0] > bmp.size[1]:
            bmp = bmp.rotate(90, expand=True)

        ratios = [1.0 * printable_area[0] / bmp.size[0], 1.0 * printable_area[0] / bmp.size[1]]
        scale = min(ratios)

        # Start the print job, and draw the bitmap to
        #  the printer device at the scaled size.

        hDC.StartDoc(file_name)
        hDC.StartPage()

        dib = ImageWin.Dib(bmp)
        scaled_width, scaled_height = [int(scale * i) for i in bmp.size]

        x1 = int((printer_size[0] - scaled_width) / 2)
        y1 = int((printer_size[1] - scaled_height) / 2)
        x2 = x1 + scaled_width
        y2 = y1 + scaled_height

        # print(f'x1 = {x1} y1 = {y1} x2= {x2} y2 = {y2}')
        dib.draw(hDC.GetHandleOutput(), (1550, 150, 3250, 5035))
        hDC.EndPage()
        hDC.EndDoc()
        hDC.DeleteDC()
        os.remove(name.value + ".jpg")


def alhabib_Chq(value):
    now = datetime.now()
    dates = str(now.strftime("%d-%m-%Y"))
    datex = dates.replace("-", "")
    datez = datex[:-2]

    chequeWidth = 811
    chequeHeight = 278

    chequeFont = 'arial.ttf'
    chequeFontSize = 20

    chequeDatex = 650
    chequeDatey = 60.8

    chequeNamex = 220
    chequeNamey = 99.28

    chequeAmountEnglishx = 193
    chequeAmountEnglishy = 125.2

    chequeAmountNumberx = 680
    chequeAmountNumbery = 122.4

    chequeFont = 'arial.ttf'
    chequeFontSize = 12

    counterDatex = 80.4
    counterDatey = 86.2

    counterNamex = 76.0
    counterNamey = 99.8

    counterAmountx = 95
    counterAmounty = 191.6

    if config == True:
        with open('alhabibConfig.ini') as f:
            for line in f:
                # For Python3, use print(line)
                if 'counterAmountx' in line:
                    txt = str(line)
                    receipts = str(re.findall(r'\d+(?:\.\d+)?', txt))
                    game = listToStringWithoutBrackets(receipts)
                    counterAmountx = float(game)
                if 'counterAmounty' in line:
                    txt = str(line)
                    receipts = str(re.findall(r'\d+(?:\.\d+)?', txt))
                    game = listToStringWithoutBrackets(receipts)
                    counterAmounty = float(game)
                if 'counterNamex' in line:
                    txt = str(line)
                    receipts = str(re.findall(r'\d+(?:\.\d+)?', txt))
                    game = listToStringWithoutBrackets(receipts)
                    counterNamex = float(game)
                if 'counterNamey' in line:
                    txt = str(line)
                    receipts = str(re.findall(r'\d+(?:\.\d+)?', txt))
                    game = listToStringWithoutBrackets(receipts)
                    counterNamey = float(game)
                if 'counterDatex' in line:
                    txt = str(line)
                    receipts = str(re.findall(r'\d+(?:\.\d+)?', txt))
                    game = listToStringWithoutBrackets(receipts)
                    counterDatex = float(game)
                if 'counterDatey' in line:
                    txt = str(line)
                    receipts = str(re.findall(r'\d+(?:\.\d+)?', txt))
                    game = listToStringWithoutBrackets(receipts)
                    counterDatey = float(game)
                if 'chequeAmountNumberx' in line:
                    txt = str(line)
                    receipts = str(re.findall(r'\d+(?:\.\d+)?', txt))
                    game = listToStringWithoutBrackets(receipts)
                    chequeAmountNumberx = float(game)
                if 'chequeAmountNumbery' in line:
                    txt = str(line)
                    receipts = str(re.findall(r'\d+(?:\.\d+)?', txt))
                    game = listToStringWithoutBrackets(receipts)
                    chequeAmountNumbery = float(game)
                if 'chequeAmountEnglishx' in line:
                    txt = str(line)
                    receipts = str(re.findall(r'\d+(?:\.\d+)?', txt))
                    game = listToStringWithoutBrackets(receipts)
                    chequeAmountEnglishx = float(game)
                if 'chequeAmountEnglishy' in line:
                    txt = str(line)
                    receipts = str(re.findall(r'\d+(?:\.\d+)?', txt))
                    game = listToStringWithoutBrackets(receipts)
                    chequeAmountEnglishy = float(game)
                if 'chequeNamex' in line:
                    txt = str(line)
                    receipts = str(re.findall(r'\d+(?:\.\d+)?', txt))
                    game = listToStringWithoutBrackets(receipts)
                    chequeNamex = float(game)
                if 'chequeNamey' in line:
                    txt = str(line)
                    receipts = str(re.findall(r'\d+(?:\.\d+)?', txt))
                    game = listToStringWithoutBrackets(receipts)
                    chequeNamey = float(game)
                if 'chequeDatex' in line:
                    txt = str(line)
                    receipts = str(re.findall(r'\d+(?:\.\d+)?', txt))
                    game = listToStringWithoutBrackets(receipts)
                    chequeDatey = float(game)
                if 'chequeDatey' in line:
                    txt = str(line)
                    receipts = str(re.findall(r'\d+(?:\.\d+)?', txt))
                    game = listToStringWithoutBrackets(receipts)
                    chequeDatey = float(game)

    wb = xl.load_workbook(value, data_only=True)
    # wb = xl.load_workbook(r'C:\Users\HYSTOU\Desktop\sest.xlsx')
    sheet = wb['Sheet1']
    cell = sheet.cell(1, 1)

    # creating image object which is of specific color
    width = 811
    height = 278

    black = 'black'
    white = 'white'

    for row in range(1, sheet.max_row + 1):
        img = PIL.Image.new(mode="RGB", size=(width, height),
                            color=(255, 255, 255))
        # img = ImageOps.expand(img, border=1, fill=black)
        draw = ImageDraw.Draw(img)
        draw.fontmode = "1"

        name = sheet.cell(row, 1)
        amount = sheet.cell(row, 2)

        # draw.line((1, 278) + img.size, fill=(0,0,0))

        # Cheque Filling

        font = PIL.ImageFont.truetype('arial.ttf', 20)
        draw.fontmode = "1"

        date = "  ".join(datez)

        draw.text((chequeDatex, chequeDatey), date + ' ' + datex[-4] + ' ' + datex[-1:], font=font, fill=black)
        draw.text((chequeNamex, chequeNamey), name.value, font=font, fill=black)
        fontSize = 20
        storeFontSize = fontSize
        wrapperWidth = 40
        caption = amount_to_million(amount.value) + " Rupees Only"

        length = len(str(caption))
        limit = 90

        if length > limit:
            calculationFontSize = int((limit * fontSize) / length)
            fontSize = calculationFontSize
            calculationWrapperWidth = (wrapperWidth * storeFontSize) / fontSize
            wrapperWidth = int(calculationWrapperWidth)

        caption1 = str(caption[0:wrapperWidth])
        caption2 = str(caption[wrapperWidth:])
        font = PIL.ImageFont.truetype("arial.ttf", fontSize)

        if len(str(caption)) > 50:
            draw.text((chequeAmountEnglishx, chequeAmountEnglishy), caption1, font=font, fill=black)
            draw.text((chequeAmountEnglishx - 55, chequeAmountEnglishy + 23), caption2, font=font, fill=black)
        font = PIL.ImageFont.truetype('arial.ttf', 20)
        payment = "{:,}".format(amount.value)
        draw.text((chequeAmountNumberx, chequeAmountNumbery), str(payment) + '/-', font=font, fill=black)

        # Counter Filling

        font = PIL.ImageFont.truetype("arial.ttf", 12)
        draw.fontmode = "1"
        draw.text((counterDatex, counterDatey), str(dates), font=font, fill=black)

        caption = name.value

        wrapper = textwrap.TextWrapper(width=16)
        word_list = wrapper.wrap(text=caption)
        caption_new = ''
        for ii in word_list[:-1]:
            caption_new = caption_new + ii + '\n'
        caption_new += word_list[-1]
        draw.text((counterNamex, counterNamey), caption_new, font=font, fill=black)
        draw.text((counterAmountx, counterAmounty), str(payment) + "/-", font=font, fill=black)

        # this will show image in any image viewer
        img.show()
        img.save(name.value + ".jpg")

        #
        # Constants for GetDeviceCaps
        #
        #
        # HORZRES / VERTRES = printable area
        #
        HORZRES = 8
        VERTRES = 10
        #
        # LOGPIXELS = dots per inch
        #
        LOGPIXELSX = 88
        LOGPIXELSY = 90
        #
        # PHYSICALWIDTH/HEIGHT = total area
        #
        PHYSICALWIDTH = 110
        PHYSICALHEIGHT = 111
        #
        # PHYSICALOFFSETX/Y = left / top margin
        #
        PHYSICALOFFSETX = 112
        PHYSICALOFFSETY = 113

        printer_name = win32print.GetDefaultPrinter()
        file_name = "test.jpg"

        #
        # You can only write a Device-independent bitmap
        #  directly to a Windows device context; therefore
        #  we need (for ease) to use the Python Imaging
        #  Library to manipulate the image.
        #
        # Create a device context from a named printer
        #  and assess the printable size of the paper.
        #
        hDC = win32ui.CreateDC()
        hDC.CreatePrinterDC(printer_name)
        printable_area = hDC.GetDeviceCaps(HORZRES), hDC.GetDeviceCaps(VERTRES)
        printer_size = hDC.GetDeviceCaps(PHYSICALWIDTH), hDC.GetDeviceCaps(PHYSICALHEIGHT)
        printer_margins = hDC.GetDeviceCaps(PHYSICALOFFSETX), hDC.GetDeviceCaps(PHYSICALOFFSETY)

        #
        # Open the image, rotate it if it's wider than
        #  it is high, and work out how much to multiply
        #  each pixel by to get it as big as possible on
        #  the page without distorting.
        #
        bmp = Image.open(name.value + ".jpg")
        if bmp.size[0] > bmp.size[1]:
            bmp = bmp.rotate(90, expand=True)

        ratios = [1.0 * printable_area[0] / bmp.size[0], 1.0 * printable_area[0] / bmp.size[1]]
        scale = min(ratios)

        # Start the print job, and draw the bitmap to
        #  the printer device at the scaled size.

        hDC.StartDoc(file_name)
        hDC.StartPage()

        dib = ImageWin.Dib(bmp)
        scaled_width, scaled_height = [int(scale * i) for i in bmp.size]

        x1 = int((printer_size[0] - scaled_width) / 2)
        y1 = int((printer_size[1] - scaled_height) / 2)
        x2 = x1 + scaled_width
        y2 = y1 + scaled_height

        # print(f'x1 = {x1} y1 = {y1} x2= {x2} y2 = {y2}')
        dib.draw(hDC.GetHandleOutput(), (1550, 150, 3250, 5035))
        hDC.EndPage()
        hDC.EndDoc()
        hDC.DeleteDC()
        os.remove(name.value + ".jpg")